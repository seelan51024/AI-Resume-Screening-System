"""
AI Resume Screening System — Backend v5
✅ HR/Manager Login Authentication (JWT)
✅ SQLite + SQLAlchemy ORM
✅ Batch-only Rankings
✅ Candidate Email with Feedback (SMTP)
✅ Excel Export (openpyxl)
✅ Role-based: HR can screen, Manager can view/export
"""

from fastapi import FastAPI, File, UploadFile, HTTPException, Depends, status, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, HTMLResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from pydantic import BaseModel, EmailStr
from typing import List, Optional
from sqlalchemy import create_engine, Column, Integer, Float, String, Boolean, DateTime, Text, func
from sqlalchemy.orm import declarative_base, sessionmaker, Session
from datetime import datetime, timedelta
from jose import JWTError, jwt
import bcrypt as _bcrypt_lib
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import pickle, json, re, os, io
import numpy as np

# ── Optional imports ───────────────────────────────────────────
try:
    import pdfplumber;  PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

try:
    import docx as python_docx;  DOCX_SUPPORT = True
except ImportError:
    DOCX_SUPPORT = False

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# ══════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════

SECRET_KEY      = "recruitai-secret-key-change-in-production-2026"
ALGORITHM       = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 480   # 8 hours

# Email config — update with real SMTP credentials
EMAIL_CONFIG = {
    "smtp_host":   "smtp.gmail.com",
    "smtp_port":   587,
    "sender_email": "your_hr_email@gmail.com",       # ← Change this
    "sender_pass":  "your_app_password_here",         # ← Gmail App Password
    "sender_name":  "RecruitAI HR System",
}

# ══════════════════════════════════════════════════════════════
# DATABASE
# ══════════════════════════════════════════════════════════════

BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
DB_PATH      = os.path.join(BASE_DIR, "resume_screener.db")
engine       = create_engine(f"sqlite:///{DB_PATH}", connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base         = declarative_base()


class UserDB(Base):
    """HR / Manager accounts"""
    __tablename__ = "users"
    id            = Column(Integer, primary_key=True, index=True)
    full_name     = Column(String(100))
    email         = Column(String(100), unique=True, index=True)
    hashed_password = Column(String(255))
    role          = Column(String(20))    # 'hr' | 'manager'
    is_active     = Column(Boolean, default=True)
    created_at    = Column(DateTime, default=datetime.utcnow)
    last_login    = Column(DateTime, nullable=True)


class CandidateDB(Base):
    """Screened resume records"""
    __tablename__ = "candidates"
    id                  = Column(Integer, primary_key=True, index=True)
    candidate_name      = Column(String(255))
    candidate_email     = Column(String(255), nullable=True)
    filename            = Column(String(255))
    job_role            = Column(String(100))
    score               = Column(Float)
    label               = Column(String(50))
    confidence          = Column(Float)
    years_experience    = Column(Integer, default=0)
    education_level     = Column(Integer, default=0)
    has_github          = Column(Boolean, default=False)
    certification_count = Column(Integer, default=0)
    gpa                 = Column(Float, default=3.0)
    internship_months   = Column(Integer, default=0)
    keyword_density     = Column(Float, default=0.0)
    skill_match_ratio   = Column(Float, default=0.0)
    num_projects        = Column(Integer, default=0)
    found_skills        = Column(Text)
    matched_skills      = Column(Text)
    missing_skills      = Column(Text)
    model_used          = Column(String(50))
    source              = Column(String(20))       # 'single' | 'batch'
    batch_session_id    = Column(String(100))
    email_sent          = Column(Boolean, default=False)
    screened_by         = Column(String(100))      # HR user email
    screened_at         = Column(DateTime, default=datetime.utcnow)
    notes               = Column(Text, nullable=True)  # HR notes/comments


class SmtpConfigDB(Base):
    """Persisted SMTP email configuration"""
    __tablename__ = "smtp_config"
    id           = Column(Integer, primary_key=True, default=1)
    smtp_host    = Column(String(255), default="smtp.gmail.com")
    smtp_port    = Column(Integer, default=587)
    sender_email = Column(String(255), default="")
    sender_pass  = Column(String(255), default="")
    sender_name  = Column(String(100), default="RecruitAI HR System")
    brevo_api_key = Column(String(255), default="")
    updated_at   = Column(DateTime, default=datetime.utcnow)


Base.metadata.create_all(bind=engine)


# ── Auto-migrate: add ALL missing columns to existing DBs ─────
def run_migrations():
    """Add any missing columns so old DBs work with new code."""
    # (table, column, sql)
    migrations = [
        ("candidates", "candidate_email",     "ALTER TABLE candidates ADD COLUMN candidate_email TEXT"),
        ("candidates", "confidence",           "ALTER TABLE candidates ADD COLUMN confidence FLOAT DEFAULT 80.0"),
        ("candidates", "years_experience",     "ALTER TABLE candidates ADD COLUMN years_experience INTEGER DEFAULT 0"),
        ("candidates", "education_level",      "ALTER TABLE candidates ADD COLUMN education_level INTEGER DEFAULT 0"),
        ("candidates", "has_github",           "ALTER TABLE candidates ADD COLUMN has_github INTEGER DEFAULT 0"),
        ("candidates", "certification_count",  "ALTER TABLE candidates ADD COLUMN certification_count INTEGER DEFAULT 0"),
        ("candidates", "gpa",                  "ALTER TABLE candidates ADD COLUMN gpa FLOAT DEFAULT 3.0"),
        ("candidates", "internship_months",    "ALTER TABLE candidates ADD COLUMN internship_months INTEGER DEFAULT 0"),
        ("candidates", "keyword_density",      "ALTER TABLE candidates ADD COLUMN keyword_density FLOAT DEFAULT 0.0"),
        ("candidates", "skill_match_ratio",    "ALTER TABLE candidates ADD COLUMN skill_match_ratio FLOAT DEFAULT 0.0"),
        ("candidates", "num_projects",         "ALTER TABLE candidates ADD COLUMN num_projects INTEGER DEFAULT 0"),
        ("candidates", "found_skills",         "ALTER TABLE candidates ADD COLUMN found_skills TEXT DEFAULT '[]'"),
        ("candidates", "matched_skills",       "ALTER TABLE candidates ADD COLUMN matched_skills TEXT DEFAULT '[]'"),
        ("candidates", "missing_skills",       "ALTER TABLE candidates ADD COLUMN missing_skills TEXT DEFAULT '[]'"),
        ("candidates", "model_used",           "ALTER TABLE candidates ADD COLUMN model_used TEXT DEFAULT 'random_forest'"),
        ("candidates", "source",               "ALTER TABLE candidates ADD COLUMN source TEXT DEFAULT 'single'"),
        ("candidates", "batch_session_id",     "ALTER TABLE candidates ADD COLUMN batch_session_id TEXT"),
        ("candidates", "email_sent",           "ALTER TABLE candidates ADD COLUMN email_sent INTEGER DEFAULT 0"),
        ("candidates", "screened_by",          "ALTER TABLE candidates ADD COLUMN screened_by TEXT"),
        ("candidates", "screened_at",          "ALTER TABLE candidates ADD COLUMN screened_at TEXT"),
        ("candidates", "notes",                "ALTER TABLE candidates ADD COLUMN notes TEXT"),
        ("users",      "full_name",            "ALTER TABLE users ADD COLUMN full_name TEXT DEFAULT ''"),
        ("users",      "is_active",            "ALTER TABLE users ADD COLUMN is_active INTEGER DEFAULT 1"),
        ("users",      "created_at",           "ALTER TABLE users ADD COLUMN created_at TEXT"),
        ("users",      "last_login",           "ALTER TABLE users ADD COLUMN last_login TEXT"),
        ("smtp_config","updated_at",           "ALTER TABLE smtp_config ADD COLUMN updated_at TEXT"),
        ("smtp_config","sender_name",          "ALTER TABLE smtp_config ADD COLUMN sender_name TEXT DEFAULT 'RecruitAI HR System'"),
        ("smtp_config","brevo_api_key",         "ALTER TABLE smtp_config ADD COLUMN brevo_api_key TEXT DEFAULT ''"),
    ]
    import sqlalchemy as _sa
    try:
        with engine.connect() as conn:
            for table, column, sql in migrations:
                try:
                    conn.execute(_sa.text(sql))
                    conn.commit()
                    print(f"  ✅ migrated: {table}.{column}")
                except Exception:
                    pass  # column already exists — skip silently
        print("✅ DB migration complete")
    except Exception as e:
        print(f"⚠️  Migration warning (non-fatal): {e}")

run_migrations()


# ══════════════════════════════════════════════════════════════
# AUTH SETUP
# ══════════════════════════════════════════════════════════════

oauth2   = OAuth2PasswordBearer(tokenUrl="auth/login")


def hash_password(pw: str) -> str:
    salt = _bcrypt_lib.gensalt()
    return _bcrypt_lib.hashpw(pw.encode("utf-8"), salt).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    try:
        return _bcrypt_lib.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False

def create_token(data: dict) -> str:
    to_encode = data.copy()
    to_encode["exp"] = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def get_current_user(token: str = Depends(oauth2), db: Session = Depends(get_db)) -> UserDB:
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("sub")
        if not email:
            raise HTTPException(status_code=401, detail="Invalid token")
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    user = db.query(UserDB).filter(UserDB.email == email).first()
    if not user or not user.is_active:
        raise HTTPException(status_code=401, detail="User not found or inactive")
    return user

def require_hr(user: UserDB = Depends(get_current_user)) -> UserDB:
    """Only HR can screen resumes"""
    if user.role not in ("hr", "manager"):
        raise HTTPException(status_code=403, detail="HR access required")
    return user

def require_manager(user: UserDB = Depends(get_current_user)) -> UserDB:
    """Manager-only endpoints"""
    if user.role != "manager":
        raise HTTPException(status_code=403, detail="Manager access required")
    return user


# ══════════════════════════════════════════════════════════════
# SEED DEFAULT USERS (run once)
# ══════════════════════════════════════════════════════════════

def seed_users():
    try:
        db = SessionLocal()
        defaults = [
            {"full_name": "HR Admin",      "email": "hr@company.com",      "password": "hr123456",  "role": "hr"},
            {"full_name": "Sarah Manager", "email": "manager@company.com",  "password": "mgr123456", "role": "manager"},
        ]
        for u in defaults:
            if not db.query(UserDB).filter(UserDB.email == u["email"]).first():
                db.add(UserDB(
                    full_name=u["full_name"], email=u["email"],
                    hashed_password=hash_password(u["password"]), role=u["role"]
                ))
        db.commit()
        db.close()
        print("✅ Default users seeded")
    except Exception as e:
        print(f"⚠️  seed_users error (non-fatal): {e}")

seed_users()

# ══════════════════════════════════════════════════════════════
# FASTAPI APP
# ══════════════════════════════════════════════════════════════

app = FastAPI(title="RecruitAI — HR Resume Screener", version="5.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# ── Serve frontend (index.html) ────────────────────────────────
FRONTEND_PATH = os.path.join(BASE_DIR, "index.html")

@app.get("/", response_class=HTMLResponse, include_in_schema=False)
def serve_frontend():
    if os.path.exists(FRONTEND_PATH):
        with open(FRONTEND_PATH, "r", encoding="utf-8") as f:
            return HTMLResponse(content=f.read())
    return HTMLResponse(content="<h1>index.html not found</h1>", status_code=404)

# ══════════════════════════════════════════════════════════════
# ML MODELS
# ══════════════════════════════════════════════════════════════

MODEL_DIR = os.path.join(BASE_DIR, "models")

def load_models():
    m = {}
    try:
        with open(os.path.join(MODEL_DIR, "random_forest.pkl"),       "rb") as f: m["random_forest"]      = pickle.load(f)
        with open(os.path.join(MODEL_DIR, "decision_tree.pkl"),        "rb") as f: m["decision_tree"]       = pickle.load(f)
        with open(os.path.join(MODEL_DIR, "label_encoder.pkl"),        "rb") as f: m["label_encoder"]       = pickle.load(f)
        with open(os.path.join(MODEL_DIR, "feature_importances.json"), "r")  as f: m["feature_importances"] = json.load(f)
        print("✅ ML Models loaded")
    except FileNotFoundError:
        print("⚠️  Models not found — run ml/train_model.py first")
    return m

MODELS = load_models()

# ══════════════════════════════════════════════════════════════
# SKILLS & JOB PROFILES
# ══════════════════════════════════════════════════════════════

TECH_SKILLS = [
    "python","java","javascript","typescript","c++","c#","go","rust","kotlin","swift",
    "react","angular","vue","next.js","node.js","django","fastapi","flask","spring boot",
    "sql","mysql","postgresql","mongodb","redis","elasticsearch","firebase",
    "aws","azure","gcp","docker","kubernetes","terraform","jenkins","github actions",
    "machine learning","deep learning","tensorflow","pytorch","scikit-learn","pandas","numpy",
    "git","linux","rest api","graphql","microservices","agile","scrum","data structures",
    "algorithms","system design","ci/cd","devops","html","css","tailwind",
]

JOB_PROFILES = {
    "Software Engineer":    ["python","java","javascript","git","data structures","algorithms","sql"],
    "Frontend Developer":   ["javascript","react","angular","vue","html","css","typescript"],
    "Backend Developer":    ["python","java","node.js","sql","rest api","docker","postgresql"],
    "Full Stack Developer": ["javascript","react","node.js","sql","html","css","git"],
    "Data Scientist":       ["python","machine learning","pandas","numpy","sql","tensorflow","scikit-learn"],
    "DevOps Engineer":      ["docker","kubernetes","aws","ci/cd","linux","terraform","jenkins"],
    "ML Engineer":          ["python","machine learning","deep learning","tensorflow","pytorch","scikit-learn"],
}

EDU_LABEL = {0: "None", 1: "Diploma", 2: "Bachelor", 3: "Master/PhD"}

# ══════════════════════════════════════════════════════════════
# RESUME PARSING
# ══════════════════════════════════════════════════════════════

def extract_text(content: bytes, filename: str) -> str:
    ext = filename.lower().rsplit(".", 1)[-1]
    if ext == "pdf" and PDF_SUPPORT:
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            return "\n".join(p.extract_text() or "" for p in pdf.pages)
    if ext in ("docx","doc") and DOCX_SUPPORT:
        d = python_docx.Document(io.BytesIO(content))
        return "\n".join(p.text for p in d.paragraphs)
    return content.decode("utf-8", errors="ignore")


def parse_resume(text: str, job_role: str) -> dict:
    t          = text.lower()
    job_skills = JOB_PROFILES.get(job_role, JOB_PROFILES["Software Engineer"])
    found      = [s for s in TECH_SKILLS if s in t]
    matched    = [s for s in job_skills if s in t]
    missing    = [s for s in job_skills if s not in t]

    exp_m  = re.findall(r'(\d+)\+?\s*years?\s*(?:of\s*)?(?:experience|exp)', t)
    years  = max((int(x) for x in exp_m), default=0)
    if not years:
        yrs = re.findall(r'20\d{2}', text)
        if len(yrs) >= 2:
            years = max(0, int(max(yrs)) - int(min(yrs)))

    edu = 0
    for lvl, kws in {3:["phd","ph.d","m.tech","m.sc","master","mba"],
                     2:["b.tech","b.sc","bachelor","b.e","bca"],
                     1:["diploma","associate"]}.items():
        if any(k in t for k in kws): edu = lvl; break

    projs  = min(len(re.findall(r'project', t)), 10)
    github = 1 if "github" in t else 0
    certs  = min(sum(1 for k in ["certified","certification","certificate","coursera","udemy"] if k in t), 5)
    gpa_m  = re.search(r'(?:gpa|cgpa)[:\s]*([0-9]\.[0-9]+)', t)
    gpa    = float(gpa_m.group(1)) if gpa_m else 3.0
    gpa    = min(4.0, max(0.0, gpa if gpa <= 4.0 else gpa / 10 * 4))
    intern = min(len(re.findall(r'intern', t)) * 3, 24)
    kd     = min(1.0, sum(t.count(s) for s in found) / max(len(text.split()), 1))
    smr    = len(matched) / max(len(job_skills), 1)

    return {
        "years_experience":    min(years, 15),
        "skill_match_ratio":   round(smr, 2),
        "education_level":     edu,
        "num_projects":        projs,
        "has_github":          github,
        "certification_count": certs,
        "gpa":                 round(gpa, 2),
        "internship_months":   intern,
        "keyword_density":     round(kd, 4),
        "found_skills":        found,
        "matched_skills":      matched,
        "missing_skills":      missing,
    }


def compute_score(f: dict) -> float:
    s = (
        min(f["years_experience"], 10) * 4 +
        f["skill_match_ratio"] * 30 +
        f["education_level"] * 6 +
        min(f["num_projects"], 8) * 2.5 +
        f["has_github"] * 5 +
        min(f["certification_count"], 4) * 2.5 +
        (f["gpa"] - 2.0) * 5 +
        min(f["internship_months"], 12) * 0.5 +
        f["keyword_density"] * 10
    )
    return round(min(100, max(0, s)), 1)


def ml_predict(f: dict, model_name: str):
    if model_name not in MODELS:
        return None, None
    model  = MODELS[model_name]
    le     = MODELS["label_encoder"]
    order  = ["years_experience","skill_match_ratio","education_level",
              "num_projects","has_github","certification_count",
              "gpa","internship_months","keyword_density"]
    X      = np.array([[f[k] for k in order]])
    enc    = model.predict(X)[0]
    proba  = model.predict_proba(X)[0]
    return le.inverse_transform([enc])[0], round(float(max(proba)) * 100, 1)


# ══════════════════════════════════════════════════════════════
# EMAIL — Candidate Feedback
# ══════════════════════════════════════════════════════════════

def build_candidate_email(candidate: dict) -> str:
    """Build professional HTML email for candidate with detailed feedback"""
    score       = candidate.get("score", 0)
    label       = candidate.get("label", "")
    name        = candidate.get("candidate_name", "Candidate")
    job_role    = candidate.get("job_role", "the applied role")
    matched     = candidate.get("matched_skills", [])
    missing     = candidate.get("missing_skills", [])
    years_exp   = candidate.get("years_experience", 0)
    edu_level   = EDU_LABEL.get(candidate.get("education_level", 0), "Not detected")
    has_github  = candidate.get("has_github", False)
    certs       = candidate.get("certification_count", 0)
    conf        = candidate.get("confidence", 0)

    # Status-specific message
    if label == "Shortlisted":
        status_color = "#1B5E20"
        status_bg    = "#E8F5E9"
        status_icon  = "✅"
        status_msg   = "Congratulations! Your profile has been <strong>shortlisted</strong> for the next round."
        next_steps   = "Our HR team will contact you within 2–3 business days to schedule an interview."
    elif label == "Under Review":
        status_color = "#E65100"
        status_bg    = "#FFF3E0"
        status_icon  = "🔍"
        status_msg   = "Your application is currently <strong>under review</strong> by our team."
        next_steps   = "We are evaluating all candidates and will update you within 5–7 business days."
    else:
        status_color = "#B71C1C"
        status_bg    = "#FFEBEE"
        status_icon  = "📋"
        status_msg   = "Thank you for applying. After careful review, your profile was <strong>not selected</strong> for this round."
        next_steps   = "We encourage you to apply again after strengthening the areas mentioned below."

    # Build matched skills HTML
    matched_html = "".join(
        f'<span style="background:#E3F2FD;color:#1565C0;padding:3px 10px;border-radius:12px;'
        f'font-size:12px;margin:3px;display:inline-block;font-weight:600;">{s}</span>'
        for s in matched
    ) if matched else '<span style="color:#999;">No matched skills detected</span>'

    # Build missing skills HTML with improvement tips
    missing_tips = {
        "python":           "Practice on LeetCode, HackerRank. Build 2–3 projects.",
        "javascript":       "Complete freeCodeCamp JS curriculum. Build a portfolio site.",
        "react":            "Complete the official React docs tutorial. Build a TODO app.",
        "sql":              "Practice on SQLZoo or Mode Analytics. Learn JOINs, indexes.",
        "docker":           "Take Docker's official getting-started tutorial on docs.docker.com.",
        "aws":              "Pursue AWS Cloud Practitioner certification (free tier available).",
        "machine learning": "Complete Andrew Ng's ML course on Coursera.",
        "git":              "Practice branching, merging on learngitbranching.js.org.",
        "data structures":  "Solve 50 LeetCode Easy/Medium problems. Study arrays, trees, graphs.",
        "algorithms":       "Study Big-O notation. Practice on Codeforces or LeetCode.",
    }

    missing_html = ""
    if missing:
        for s in missing[:6]:
            tip = missing_tips.get(s, f"Study and practice {s} through online courses and projects.")
            missing_html += (
                f'<tr>'
                f'<td style="padding:10px 12px;border-bottom:1px solid #f0f0f0;">'
                f'<span style="background:#FFEBEE;color:#B71C1C;padding:2px 8px;border-radius:8px;'
                f'font-size:12px;font-weight:600;">❌ {s}</span></td>'
                f'<td style="padding:10px 12px;border-bottom:1px solid #f0f0f0;font-size:13px;color:#444;">'
                f'{tip}</td>'
                f'</tr>'
            )
    else:
        missing_html = '<tr><td colspan="2" style="padding:12px;color:#666;">No critical missing skills!</td></tr>'

    # Score bar width
    bar_width = int(score)
    bar_color = "#1B5E20" if score >= 70 else "#E65100" if score >= 45 else "#B71C1C"

    html = f"""
<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#F8FAFF;font-family:'Segoe UI',Arial,sans-serif;">

<div style="max-width:640px;margin:30px auto;background:#ffffff;border-radius:16px;
            box-shadow:0 4px 24px rgba(10,22,40,0.12);overflow:hidden;">

  <!-- Header -->
  <div style="background:linear-gradient(135deg,#0a1628,#1565C0);padding:32px 36px;">
    <div style="display:flex;align-items:center;gap:12px;">
      <div style="background:rgba(255,255,255,0.15);border-radius:10px;padding:8px 14px;
                  font-size:22px;">🤖</div>
      <div>
        <div style="color:#fff;font-size:22px;font-weight:800;letter-spacing:-0.5px;">RecruitAI</div>
        <div style="color:#90CAF9;font-size:12px;font-weight:500;letter-spacing:1px;">HR SCREENING SYSTEM</div>
      </div>
    </div>
  </div>

  <!-- Status Banner -->
  <div style="background:{status_bg};border-bottom:3px solid {status_color};padding:20px 36px;">
    <div style="font-size:18px;color:{status_color};font-weight:700;">{status_icon} Application Status Update</div>
    <div style="font-size:14px;color:#444;margin-top:6px;line-height:1.6;">{status_msg}</div>
  </div>

  <!-- Body -->
  <div style="padding:32px 36px;">

    <p style="font-size:16px;color:#0a1628;font-weight:600;margin-bottom:4px;">Dear {name},</p>
    <p style="font-size:14px;color:#555;margin-bottom:24px;line-height:1.7;">
      Thank you for applying for the <strong>{job_role}</strong> position.
      Our AI-powered screening system has completed the analysis of your resume.
      Here is your detailed evaluation report:
    </p>

    <!-- Score Card -->
    <div style="background:#F8FAFF;border:1.5px solid #BBDEFB;border-radius:12px;
                padding:20px 24px;margin-bottom:24px;">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;">
        <div style="font-size:13px;font-weight:700;color:#546E7A;text-transform:uppercase;
                    letter-spacing:0.08em;">AI Screening Score</div>
        <div style="font-size:28px;font-weight:800;color:{bar_color};">{score}<span style="font-size:14px;color:#999;">/100</span></div>
      </div>
      <div style="height:10px;background:#E3F2FD;border-radius:6px;overflow:hidden;">
        <div style="width:{bar_width}%;height:100%;background:{bar_color};border-radius:6px;"></div>
      </div>
      <div style="display:flex;justify-content:space-between;margin-top:8px;">
        <span style="font-size:11px;color:#999;">0</span>
        <span style="font-size:11px;color:#999;font-weight:600;">Model Confidence: {conf}%</span>
        <span style="font-size:11px;color:#999;">100</span>
      </div>
    </div>

    <!-- Profile Summary -->
    <div style="margin-bottom:24px;">
      <div style="font-size:14px;font-weight:700;color:#0a1628;margin-bottom:12px;">
        📊 Profile Summary
      </div>
      <table style="width:100%;border-collapse:collapse;">
        <tr style="background:#F8FAFF;">
          <td style="padding:8px 12px;font-size:13px;color:#546E7A;font-weight:600;width:40%;">Experience</td>
          <td style="padding:8px 12px;font-size:13px;color:#0a1628;font-weight:700;">{years_exp} year(s)</td>
        </tr>
        <tr>
          <td style="padding:8px 12px;font-size:13px;color:#546E7A;font-weight:600;">Education</td>
          <td style="padding:8px 12px;font-size:13px;color:#0a1628;font-weight:700;">{edu_level}</td>
        </tr>
        <tr style="background:#F8FAFF;">
          <td style="padding:8px 12px;font-size:13px;color:#546E7A;font-weight:600;">GitHub Profile</td>
          <td style="padding:8px 12px;font-size:13px;color:#0a1628;font-weight:700;">{'✅ Detected' if has_github else '❌ Not found — Add GitHub link to resume'}</td>
        </tr>
        <tr>
          <td style="padding:8px 12px;font-size:13px;color:#546E7A;font-weight:600;">Certifications</td>
          <td style="padding:8px 12px;font-size:13px;color:#0a1628;font-weight:700;">{certs} detected</td>
        </tr>
      </table>
    </div>

    <!-- Matched Skills -->
    <div style="margin-bottom:24px;">
      <div style="font-size:14px;font-weight:700;color:#0a1628;margin-bottom:10px;">
        ✅ Skills You Have (Matched with Job Requirements)
      </div>
      <div style="line-height:2;">{matched_html}</div>
    </div>

    <!-- Missing Skills with Tips -->
    <div style="margin-bottom:24px;">
      <div style="font-size:14px;font-weight:700;color:#0a1628;margin-bottom:10px;">
        📈 Areas to Strengthen (Missing Skills & Improvement Tips)
      </div>
      <table style="width:100%;border-collapse:collapse;border:1px solid #f0f0f0;border-radius:8px;overflow:hidden;">
        <tr style="background:#0a1628;">
          <th style="padding:10px 12px;text-align:left;font-size:12px;color:#90CAF9;font-weight:600;">Missing Skill</th>
          <th style="padding:10px 12px;text-align:left;font-size:12px;color:#90CAF9;font-weight:600;">How to Improve</th>
        </tr>
        {missing_html}
      </table>
    </div>

    <!-- Next Steps -->
    <div style="background:linear-gradient(135deg,#E3F2FD,#EFF6FF);border-left:4px solid #1565C0;
                border-radius:8px;padding:16px 20px;margin-bottom:24px;">
      <div style="font-size:13px;font-weight:700;color:#0D47A1;margin-bottom:6px;">📌 Next Steps</div>
      <div style="font-size:13px;color:#444;line-height:1.6;">{next_steps}</div>
    </div>

    <p style="font-size:13px;color:#777;line-height:1.7;">
      This evaluation was performed by our AI screening system after initial ATS filtering.
      Scores are based on skill match, experience, education, projects, and other profile attributes.
      We wish you the best in your career journey!
    </p>

  </div>

  <!-- Footer -->
  <div style="background:#F8FAFF;border-top:1px solid #E3F2FD;padding:20px 36px;text-align:center;">
    <div style="font-size:12px;color:#999;line-height:1.8;">
      This is an automated message from the RecruitAI HR Screening System.<br>
      Please do not reply to this email. For queries, contact hr@company.com<br>
      <span style="color:#BBDEFB;">─────────────────────────────</span><br>
      © {datetime.now().year} RecruitAI · Confidential HR Communication
    </div>
  </div>

</div>
</body>
</html>
"""
    return html


def get_smtp_config(db=None) -> dict:
    """Read SMTP config from DB, fallback to static EMAIL_CONFIG"""
    if db:
        row = db.query(SmtpConfigDB).filter(SmtpConfigDB.id == 1).first()
        if row and row.sender_email:
            return {
                "smtp_host":     row.smtp_host,
                "smtp_port":     row.smtp_port,
                "sender_email":  row.sender_email,
                "sender_pass":   row.sender_pass,
                "sender_name":   row.sender_name,
                "brevo_api_key": getattr(row, "brevo_api_key", "") or "",
            }
    return EMAIL_CONFIG


def send_via_brevo_api(to_email: str, candidate: dict, cfg: dict) -> bool:
    """Send email using Brevo HTTP API — works even when SMTP ports are blocked"""
    import urllib.request, urllib.error, json as _json
    api_key = cfg.get("brevo_api_key", "")
    if not api_key:
        return False
    sender_email = cfg.get("sender_email", "")
    sender_name  = cfg.get("sender_name", "RecruitAI HR System")
    html_body    = build_candidate_email(candidate)
    subject      = f"Your Application Status — {candidate.get('job_role','Position')} | RecruitAI"
    payload = _json.dumps({
        "sender":      {"name": sender_name, "email": sender_email},
        "to":          [{"email": to_email}],
        "subject":     subject,
        "htmlContent": html_body,
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://api.brevo.com/v3/smtp/email",
        data=payload,
        headers={"api-key": api_key, "Content-Type": "application/json", "Accept": "application/json"},
        method="POST"
    )
    try:
        resp = urllib.request.urlopen(req, timeout=20)
        result = resp.read().decode()
        print(f"✅ Brevo API email sent to {to_email}: {result}")
        return True, None
    except urllib.error.HTTPError as e:
        err_body = e.read().decode()
        print(f"Brevo API error {e.code}: {err_body}")
        if e.code == 403:
            return False, "Brevo sender not verified. Brevo dashboard → Senders & IPs → sender email-ஐ verify பண்ணுங்க."
        return False, f"Brevo API error {e.code}: {err_body}"
    except Exception as e:
        print(f"Brevo API exception: {e}")
        return False, str(e)


def send_candidate_email(to_email: str, candidate: dict, db=None):
    """Send feedback email — tries Brevo API first, falls back to SMTP. Returns (success, error_msg)."""
    try:
        cfg = get_smtp_config(db)
        if not cfg.get("sender_email") or cfg["sender_email"] == "your_hr_email@gmail.com":
            print("Email not sent: not configured.")
            return False, "SMTP not configured"

        # ── Try Brevo API first (works even when SMTP ports are blocked) ──
        if cfg.get("brevo_api_key"):
            print("Using Brevo API...")
            ok, brevo_err = send_via_brevo_api(to_email, candidate, cfg)
            if ok:
                return True, None
            print(f"Brevo API failed ({brevo_err}), falling back to SMTP...")

        # ── Fallback: SMTP ────────────────────────────────────────────────
        login_user  = cfg["sender_email"]
        login_pass  = cfg["sender_pass"]
        sender_name = cfg.get("sender_name", "RecruitAI HR System")
        html_body   = build_candidate_email(candidate)

        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"Your Application Status — {candidate.get('job_role','Position')} | RecruitAI"
        msg["From"]    = f"{sender_name} <{login_user}>"
        msg["To"]      = to_email
        msg.attach(MIMEText(html_body, "html"))

        with smtplib.SMTP(cfg["smtp_host"], int(cfg["smtp_port"])) as server:
            server.ehlo(); server.starttls(); server.ehlo()
            server.login(login_user, login_pass)
            server.sendmail(login_user, to_email, msg.as_string())
        print(f"✅ SMTP email sent to {to_email}")
        return True, None
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"Email error: {e}")
        return False, str(e)


# ══════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════════

def build_excel(candidates: list, label: str = "Batch") -> bytes:
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Rank List"
    THIN   = Side(style="thin", color="CCCCCC")
    BDR    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT   = Alignment(horizontal="left",  vertical="center", wrap_text=True)

    ws.merge_cells("A1:O1")
    c = ws["A1"]
    c.value     = f"RecruitAI — Rank List  |  {label}  |  Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    c.font      = Font(bold=True, size=13, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor="0A1628")
    c.alignment = CENTER
    ws.row_dimensions[1].height = 30

    hdrs = ["Rank","Candidate Name","Email","Job Role","AI Score","Status",
            "Confidence %","Exp (Yrs)","Education","GitHub","Certifications",
            "Matched Skills","Missing Skills","AI Model","Screened At"]
    for col, h in enumerate(hdrs, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = Font(bold=True, color="90CAF9", size=10)
        cell.fill      = PatternFill("solid", fgColor="0D2040")
        cell.alignment = CENTER
        cell.border    = BDR
    ws.row_dimensions[2].height = 22

    for i, c in enumerate(candidates):
        row   = i + 3
        rank  = i + 1
        score = c.get("score", 0)
        label_val = c.get("label", "")
        if rank == 1:   rf = PatternFill("solid", fgColor="FFD700")
        elif rank == 2: rf = PatternFill("solid", fgColor="C0C0C0")
        elif rank == 3: rf = PatternFill("solid", fgColor="CD7F32")
        elif label_val == "Shortlisted":  rf = PatternFill("solid", fgColor="D4EDDA")
        elif label_val == "Under Review": rf = PatternFill("solid", fgColor="FFF3CD")
        else:           rf = PatternFill("solid", fgColor="F8D7DA")

        vals = [rank, c.get("candidate_name",""), c.get("candidate_email","—"),
                c.get("job_role",""), score, label_val,
                c.get("confidence",0), c.get("years_experience",0),
                EDU_LABEL.get(c.get("education_level",0),"—"),
                "Yes" if c.get("has_github") else "No",
                c.get("certification_count",0),
                ", ".join(c.get("matched_skills",[])),
                ", ".join(c.get("missing_skills",[])),
                c.get("model_used","").replace("_"," ").title(),
                str(c.get("screened_at",""))[:19].replace("T"," ")]

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = rf; cell.border = BDR
            cell.alignment = CENTER if col in (1,5,6,7,8,10,11,14) else LEFT
            if col == 1:  cell.font = Font(bold=True, size=12)
            if col == 5:  cell.font = Font(bold=True, color="006400" if score>=70 else "8B4513" if score>=45 else "8B0000")
        ws.row_dimensions[row].height = 20

    widths = [7,22,24,20,12,14,13,10,14,8,13,32,32,16,20]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Summary Statistics"
    ws2["A1"].font = Font(bold=True, size=14)
    scores = [c.get("score",0) for c in candidates]
    lbs    = [c.get("label","") for c in candidates]
    for r,(k,v) in enumerate([
        ("Total Candidates", len(candidates)),
        ("Average Score",    round(sum(scores)/max(len(scores),1),1)),
        ("Highest Score",    max(scores,default=0)),
        ("Lowest Score",     min(scores,default=0)),
        ("Shortlisted ✅",   lbs.count("Shortlisted")),
        ("Under Review 🔍",  lbs.count("Under Review")),
        ("Rejected ❌",      lbs.count("Rejected")),
        ("Export Date",      datetime.now().strftime("%d-%b-%Y %H:%M")),
    ], 3):
        ws2.cell(row=r,column=1,value=k).font = Font(bold=True)
        ws2.cell(row=r,column=2,value=v)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 18

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return out.read()


# ══════════════════════════════════════════════════════════════
# HELPER
# ══════════════════════════════════════════════════════════════

def row_to_dict(c: CandidateDB) -> dict:
    return {
        "id": c.id, "candidate_name": c.candidate_name,
        "candidate_email": c.candidate_email,
        "filename": c.filename, "job_role": c.job_role,
        "score": c.score, "label": c.label, "confidence": c.confidence,
        "years_experience": c.years_experience, "education_level": c.education_level,
        "has_github": c.has_github, "certification_count": c.certification_count,
        "gpa": c.gpa, "internship_months": c.internship_months,
        "skill_match_ratio": c.skill_match_ratio, "num_projects": c.num_projects,
        "found_skills":   json.loads(c.found_skills   or "[]"),
        "matched_skills": json.loads(c.matched_skills or "[]"),
        "missing_skills": json.loads(c.missing_skills or "[]"),
        "model_used": c.model_used, "source": c.source,
        "batch_session_id": c.batch_session_id,
        "email_sent": c.email_sent,
        "screened_by": c.screened_by,
        "screened_at": c.screened_at.isoformat() if c.screened_at else "",
        "notes": c.notes or "",
    }


# ══════════════════════════════════════════════════════════════
# AUTH ROUTES
# ══════════════════════════════════════════════════════════════

class LoginResponse(BaseModel):
    access_token: str
    token_type: str
    user: dict

class RegisterRequest(BaseModel):
    full_name: str
    email: str
    password: str
    role: str   # 'hr' | 'manager'
    admin_key: str   # secret key to prevent public registration

ADMIN_REGISTER_KEY = "recruitai-admin-2026"   # Change in production

@app.post("/auth/login", response_model=LoginResponse)
def login(form: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(UserDB).filter(UserDB.email == form.username).first()
    if not user or not verify_password(form.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    if not user.is_active:
        raise HTTPException(status_code=403, detail="Account is deactivated")
    user.last_login = datetime.utcnow()
    db.commit()
    token = create_token({"sub": user.email, "role": user.role})
    return {
        "access_token": token, "token_type": "bearer",
        "user": {"id": user.id, "full_name": user.full_name,
                 "email": user.email, "role": user.role}
    }

@app.post("/auth/register")
def register(req: RegisterRequest, db: Session = Depends(get_db)):
    if req.admin_key != ADMIN_REGISTER_KEY:
        raise HTTPException(status_code=403, detail="Invalid admin key")
    if req.role not in ("hr", "manager"):
        raise HTTPException(status_code=400, detail="Role must be 'hr' or 'manager'")
    if db.query(UserDB).filter(UserDB.email == req.email).first():
        raise HTTPException(status_code=400, detail="Email already registered")
    user = UserDB(full_name=req.full_name, email=req.email,
                  hashed_password=hash_password(req.password), role=req.role)
    db.add(user); db.commit(); db.refresh(user)
    return {"message": f"User {req.email} created with role {req.role}"}

@app.get("/auth/me")
def get_me(user: UserDB = Depends(get_current_user)):
    return {"id": user.id, "full_name": user.full_name,
            "email": user.email, "role": user.role,
            "last_login": user.last_login.isoformat() if user.last_login else None}


# ══════════════════════════════════════════════════════════════
# CORE ROUTES
# ══════════════════════════════════════════════════════════════

@app.get("/health")
def health(db: Session = Depends(get_db)):
    return {
        "status": "ok", "version": "5.0.0",
        "models_loaded": bool(MODELS),
        "database": DB_PATH,
        "database_exists": os.path.exists(DB_PATH),
        "total_candidates": db.query(func.count(CandidateDB.id)).scalar(),
        "batch_candidates": db.query(func.count(CandidateDB.id)).filter(CandidateDB.source=="batch").scalar(),
    }

@app.get("/job-roles")
def job_roles(): return {"roles": list(JOB_PROFILES.keys())}


# ── Single Screen (HR only) ────────────────────────────────────
@app.post("/screen")
async def screen_resume(
    file: UploadFile = File(...),
    job_role: str = Form("Software Engineer"),
    model: str = Form("random_forest"),
    candidate_name: str = Form("Unknown"),
    candidate_email: str = Form(""),
    db: Session = Depends(get_db),
    user: UserDB = Depends(require_hr),
):
    if model not in ("random_forest","decision_tree"):
        raise HTTPException(400, "model must be random_forest or decision_tree")
    content = await file.read()
    text    = extract_text(content, file.filename)
    if len(text.strip()) < 50:
        raise HTTPException(400, "Cannot extract text — upload PDF/DOCX/TXT")
    feat  = parse_resume(text, job_role)
    score = compute_score(feat)
    label, conf = ml_predict(feat, model)
    if label is None:
        label = "Shortlisted" if score>=70 else "Under Review" if score>=45 else "Rejected"
        conf  = 75.0
    row = CandidateDB(
        candidate_name=candidate_name, candidate_email=candidate_email or None,
        filename=file.filename, job_role=job_role, score=score, label=label,
        confidence=conf, years_experience=feat["years_experience"],
        education_level=feat["education_level"], has_github=bool(feat["has_github"]),
        certification_count=feat["certification_count"], gpa=feat["gpa"],
        internship_months=feat["internship_months"], keyword_density=feat["keyword_density"],
        skill_match_ratio=feat["skill_match_ratio"], num_projects=feat["num_projects"],
        found_skills=json.dumps(feat["found_skills"]),
        matched_skills=json.dumps(feat["matched_skills"]),
        missing_skills=json.dumps(feat["missing_skills"]),
        model_used=model, source="single", screened_by=user.email,
        screened_at=datetime.utcnow(),
    )
    db.add(row); db.commit(); db.refresh(row)
    return row_to_dict(row)


# ── Batch Screen (HR only) ─────────────────────────────────────
@app.post("/screen-batch")
async def screen_batch(
    files: List[UploadFile] = File(...),
    job_role: str = Form("Software Engineer"),
    model: str = Form("random_forest"),
    db: Session = Depends(get_db),
    user: UserDB = Depends(require_hr),
):
    session_id = datetime.utcnow().strftime("BATCH_%Y%m%d_%H%M%S")
    results    = []
    for file in files:
        content = await file.read()
        text    = extract_text(content, file.filename)
        if len(text.strip()) < 10: continue
        feat  = parse_resume(text, job_role)
        score = compute_score(feat)
        label, conf = ml_predict(feat, model)
        if label is None:
            label = "Shortlisted" if score>=70 else "Under Review" if score>=45 else "Rejected"
            conf  = 75.0
        cname = file.filename.rsplit(".",1)[0]
        row = CandidateDB(
            candidate_name=cname, filename=file.filename,
            job_role=job_role, score=score, label=label, confidence=conf,
            years_experience=feat["years_experience"], education_level=feat["education_level"],
            has_github=bool(feat["has_github"]), certification_count=feat["certification_count"],
            gpa=feat["gpa"], internship_months=feat["internship_months"],
            keyword_density=feat["keyword_density"], skill_match_ratio=feat["skill_match_ratio"],
            num_projects=feat["num_projects"],
            found_skills=json.dumps(feat["found_skills"]),
            matched_skills=json.dumps(feat["matched_skills"]),
            missing_skills=json.dumps(feat["missing_skills"]),
            model_used=model, source="batch", batch_session_id=session_id,
            screened_by=user.email, screened_at=datetime.utcnow(),
        )
        db.add(row); db.commit(); db.refresh(row)
        results.append(row_to_dict(row))
    results.sort(key=lambda x: x["score"], reverse=True)
    for i,r in enumerate(results): r["rank"] = i+1
    return {"total": len(results), "batch_session_id": session_id, "results": results}


# ── Rankings (HR + Manager) ────────────────────────────────────
@app.get("/rankings")
def get_rankings(
    job_role: Optional[str] = None,
    batch_session_id: Optional[str] = None,
    label: Optional[str] = None,
    search: Optional[str] = None,
    limit: int = 200,
    db: Session = Depends(get_db),
    user: UserDB = Depends(get_current_user),
):
    q = db.query(CandidateDB).filter(CandidateDB.source=="batch")
    if job_role:          q = q.filter(CandidateDB.job_role==job_role)
    if batch_session_id:  q = q.filter(CandidateDB.batch_session_id==batch_session_id)
    if label:             q = q.filter(CandidateDB.label==label)
    if search:
        s = f"%{search}%"
        q = q.filter(
            CandidateDB.candidate_name.ilike(s) |
            CandidateDB.candidate_email.ilike(s) |
            CandidateDB.job_role.ilike(s)
        )
    rows = q.order_by(CandidateDB.score.desc()).limit(limit).all()
    data = [row_to_dict(r) for r in rows]
    for i,d in enumerate(data): d["rank"] = i+1
    return {"total": len(data), "results": data}


# ── Batch Sessions ─────────────────────────────────────────────
@app.get("/batch-sessions")
def batch_sessions(db: Session = Depends(get_db), user: UserDB = Depends(get_current_user)):
    rows = (db.query(CandidateDB.batch_session_id, CandidateDB.job_role,
                     func.count(CandidateDB.id).label("count"),
                     func.max(CandidateDB.screened_at).label("screened_at"))
            .filter(CandidateDB.source=="batch")
            .group_by(CandidateDB.batch_session_id)
            .order_by(func.max(CandidateDB.screened_at).desc()).all())
    return [{"batch_session_id": r.batch_session_id, "job_role": r.job_role,
             "count": r.count,
             "screened_at": r.screened_at.isoformat() if r.screened_at else ""} for r in rows]


# ── Send Email to Candidate ────────────────────────────────────
class EmailRequest(BaseModel):
    candidate_id: int
    to_email: str
    mailto_sent: bool = False  # True when sent via mailto (no SMTP needed)

@app.post("/send-email")
def send_email_to_candidate(
    req: EmailRequest,
    db: Session = Depends(get_db),
    user: UserDB = Depends(require_hr),
):
    cand = db.query(CandidateDB).filter(CandidateDB.id == req.candidate_id).first()
    if not cand:
        raise HTTPException(404, "Candidate not found")

    # mailto mode — just mark as sent, no SMTP needed
    if req.mailto_sent:
        cand.email_sent = True
        cand.candidate_email = req.to_email
        db.commit()
        return {"message": f"Marked as sent to {req.to_email}", "success": True}

    # SMTP mode — check config first
    cfg = get_smtp_config(db)
    if not cfg.get("sender_email") or cfg["sender_email"] == "your_hr_email@gmail.com":
        raise HTTPException(400, "SMTP not configured. Go to Settings → Email Config to set up Gmail SMTP.")

    # Always save email in DB before sending
    cand.candidate_email = req.to_email
    db.commit()

    data    = row_to_dict(cand)
    success, err_msg = send_candidate_email(req.to_email, data, db=db)
    if success:
        cand.email_sent = True
        db.commit()
        return {"message": f"Email sent successfully to {req.to_email}", "success": True}
    else:
        detail = err_msg or "Email delivery failed — verify SMTP/Brevo settings"
        raise HTTPException(500, detail)


# ── Preview Email HTML ─────────────────────────────────────────
@app.get("/email-preview/{candidate_id}")
def preview_email(
    candidate_id: int,
    db: Session = Depends(get_db),
    user: UserDB = Depends(require_hr),
):
    cand = db.query(CandidateDB).filter(CandidateDB.id == candidate_id).first()
    if not cand:
        raise HTTPException(404, "Candidate not found")
    from fastapi.responses import HTMLResponse
    return HTMLResponse(content=build_candidate_email(row_to_dict(cand)))


# ── Excel Export (HR + Manager) ────────────────────────────────
@app.get("/export/excel")
def export_excel(
    job_role: Optional[str] = None,
    batch_session_id: Optional[str] = None,
    db: Session = Depends(get_db),
    user: UserDB = Depends(get_current_user),
):
    if not EXCEL_SUPPORT:
        raise HTTPException(500, "Install openpyxl: pip install openpyxl")
    q = db.query(CandidateDB).filter(CandidateDB.source=="batch")
    if job_role:          q = q.filter(CandidateDB.job_role==job_role)
    if batch_session_id:  q = q.filter(CandidateDB.batch_session_id==batch_session_id)
    rows = q.order_by(CandidateDB.score.desc()).all()
    if not rows: raise HTTPException(404, "No batch candidates to export")
    data = [row_to_dict(r) for r in rows]
    for i,c in enumerate(data): c["rank"] = i+1
    lbl     = batch_session_id or job_role or "All Batches"
    content = build_excel(data, lbl)
    fname   = f"RecruitAI_RankList_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ── Stats ──────────────────────────────────────────────────────
@app.get("/stats")
def stats(db: Session = Depends(get_db), user: UserDB = Depends(get_current_user)):
    total = db.query(func.count(CandidateDB.id)).scalar()
    if not total: return {"message": "No candidates screened yet"}
    return {
        "total_screened":  total,
        "batch_screened":  db.query(func.count(CandidateDB.id)).filter(CandidateDB.source=="batch").scalar(),
        "single_screened": db.query(func.count(CandidateDB.id)).filter(CandidateDB.source=="single").scalar(),
        "average_score":   round(db.query(func.avg(CandidateDB.score)).scalar() or 0, 1),
        "highest_score":   round(db.query(func.max(CandidateDB.score)).scalar() or 0, 1),
        "lowest_score":    round(db.query(func.min(CandidateDB.score)).scalar() or 0, 1),
        "shortlisted":     db.query(func.count(CandidateDB.id)).filter(CandidateDB.label=="Shortlisted").scalar(),
        "under_review":    db.query(func.count(CandidateDB.id)).filter(CandidateDB.label=="Under Review").scalar(),
        "rejected":        db.query(func.count(CandidateDB.id)).filter(CandidateDB.label=="Rejected").scalar(),
        "emails_sent":     db.query(func.count(CandidateDB.id)).filter(CandidateDB.email_sent==True).scalar(),
    }

@app.get("/feature-importances")
def feature_importances(user: UserDB = Depends(get_current_user)):
    if "feature_importances" not in MODELS:
        raise HTTPException(404, "Train models first")
    return MODELS["feature_importances"]

@app.delete("/results")
def clear_results(db: Session = Depends(get_db), user: UserDB = Depends(require_manager)):
    n = db.query(CandidateDB).delete(); db.commit()
    return {"message": f"Cleared {n} records"}

# ══════════════════════════════════════════════════════════════
# NEW FEATURES
# ══════════════════════════════════════════════════════════════

# ── SMTP Config (read/write from DB) ──────────────────────────
class SmtpConfigRequest(BaseModel):
    smtp_host: str = "smtp.gmail.com"
    smtp_port: int = 587
    sender_email: str
    sender_pass: str = ""  # Optional — blank means "keep existing password"
    sender_name: str = "RecruitAI HR System"
    brevo_api_key: str = ""  # Optional — Brevo API key for reliable sending

@app.get("/smtp-config")
def get_smtp_config_route(db: Session = Depends(get_db), user: UserDB = Depends(get_current_user)):
    row = db.query(SmtpConfigDB).filter(SmtpConfigDB.id == 1).first()
    if not row:
        return {"smtp_host": "smtp.gmail.com", "smtp_port": 587, "sender_email": "",
                "sender_name": "RecruitAI HR System", "configured": False}
    return {
        "smtp_host": row.smtp_host, "smtp_port": row.smtp_port,
        "sender_email": row.sender_email, "sender_name": row.sender_name,
        "configured": bool(row.sender_email),
    }

@app.post("/smtp-config")
def save_smtp_config(req: SmtpConfigRequest, db: Session = Depends(get_db),
                     user: UserDB = Depends(require_manager)):
    row = db.query(SmtpConfigDB).filter(SmtpConfigDB.id == 1).first()
    if row:
        row.smtp_host    = req.smtp_host
        row.smtp_port    = req.smtp_port
        row.sender_email = req.sender_email
        row.sender_name  = req.sender_name
        row.updated_at   = datetime.utcnow()
        if req.sender_pass:  # Only overwrite password if a new one is provided
            row.sender_pass = req.sender_pass
        if req.brevo_api_key:  # Only overwrite API key if a new one is provided
            row.brevo_api_key = req.brevo_api_key
    else:
        db.add(SmtpConfigDB(id=1, smtp_host=req.smtp_host, smtp_port=req.smtp_port,
                            sender_email=req.sender_email, sender_pass=req.sender_pass,
                            sender_name=req.sender_name, brevo_api_key=req.brevo_api_key))
    db.commit()
    return {"message": "SMTP config saved", "sender_email": req.sender_email}

@app.post("/smtp-test")
def test_smtp(db: Session = Depends(get_db), user: UserDB = Depends(require_manager)):
    """Send a test email — tries Brevo API first, then SMTP"""
    import urllib.request, urllib.error, json as _json
    cfg = get_smtp_config(db)
    if not cfg.get("sender_email") or cfg["sender_email"] == "your_hr_email@gmail.com":
        raise HTTPException(400, "Not configured. Save config first.")

    # ── Try Brevo API ────────────────────────────────────────────
    if cfg.get("brevo_api_key"):
        payload = _json.dumps({
            "sender":      {"name": cfg.get("sender_name","RecruitAI"), "email": cfg["sender_email"]},
            "to":          [{"email": cfg["sender_email"]}],
            "subject":     "✅ RecruitAI — Email Config Test Successful",
            "htmlContent": "<h2>🎉 Brevo API is working!</h2><p>Your RecruitAI email config is correct.</p>",
        }).encode("utf-8")
        req = urllib.request.Request(
            "https://api.brevo.com/v3/smtp/email",
            data=payload,
            headers={"api-key": cfg["brevo_api_key"], "Content-Type": "application/json"},
            method="POST"
        )
        try:
            urllib.request.urlopen(req, timeout=20)
            return {"success": True, "message": f"✅ Test email sent via Brevo API to {cfg['sender_email']}"}
        except urllib.error.HTTPError as e:
            err = e.read().decode()
            raise HTTPException(400, f"Brevo API error {e.code}: {err}")
        except Exception as e:
            raise HTTPException(400, f"Brevo API error: {str(e)}")

    # ── Fallback SMTP ────────────────────────────────────────────
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = "✅ RecruitAI — SMTP Test Successful"
        msg["From"]    = f"{cfg['sender_name']} <{cfg['sender_email']}>"
        msg["To"]      = cfg["sender_email"]
        msg.attach(MIMEText("<h2>SMTP is working! 🎉</h2><p>Your RecruitAI email config is correct.</p>", "html"))
        with smtplib.SMTP(cfg["smtp_host"], int(cfg["smtp_port"])) as server:
            server.ehlo(); server.starttls(); server.ehlo()
            server.login(cfg["sender_email"], cfg["sender_pass"])
            server.sendmail(cfg["sender_email"], cfg["sender_email"], msg.as_string())
        return {"success": True, "message": f"✅ Test email sent via SMTP to {cfg['sender_email']}"}
    except Exception as e:
        raise HTTPException(400, f"SMTP error: {str(e)}")


# ── Candidate Notes ────────────────────────────────────────────
class NoteRequest(BaseModel):
    notes: str

@app.patch("/candidates/{candidate_id}/notes")
def update_notes(candidate_id: int, req: NoteRequest,
                 db: Session = Depends(get_db), user: UserDB = Depends(require_hr)):
    cand = db.query(CandidateDB).filter(CandidateDB.id == candidate_id).first()
    if not cand:
        raise HTTPException(404, "Candidate not found")
    cand.notes = req.notes
    db.commit()
    return {"message": "Notes saved", "candidate_id": candidate_id}


# ── Candidate Label Override (Manager only) ────────────────────
class LabelRequest(BaseModel):
    label: str   # Shortlisted | Under Review | Rejected

@app.patch("/candidates/{candidate_id}/label")
def update_label(candidate_id: int, req: LabelRequest,
                 db: Session = Depends(get_db), user: UserDB = Depends(require_manager)):
    if req.label not in ("Shortlisted", "Under Review", "Rejected"):
        raise HTTPException(400, "Invalid label")
    cand = db.query(CandidateDB).filter(CandidateDB.id == candidate_id).first()
    if not cand:
        raise HTTPException(404, "Candidate not found")
    cand.label = req.label
    db.commit()
    return {"message": f"Label updated to {req.label}", "candidate_id": candidate_id}


# ── All Candidates (Single + Batch) ───────────────────────────
@app.get("/candidates")
def all_candidates(
    search: Optional[str] = None,
    label: Optional[str] = None,
    job_role: Optional[str] = None,
    limit: int = 200,
    db: Session = Depends(get_db),
    user: UserDB = Depends(get_current_user),
):
    q = db.query(CandidateDB)
    if label:    q = q.filter(CandidateDB.label == label)
    if job_role: q = q.filter(CandidateDB.job_role == job_role)
    if search:
        s = f"%{search}%"
        q = q.filter(
            CandidateDB.candidate_name.ilike(s) |
            CandidateDB.candidate_email.ilike(s)
        )
    rows = q.order_by(CandidateDB.score.desc()).limit(limit).all()
    data = [row_to_dict(r) for r in rows]
    for i, d in enumerate(data): d["rank"] = i + 1
    return {"total": len(data), "results": data}
